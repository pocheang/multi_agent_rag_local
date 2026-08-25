"""Unified DOCX, PPTX, XLSX, and XLS evidence parsing."""

from __future__ import annotations

import hashlib
import mimetypes
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from langchain_core.documents import Document

from app.services.evidence.models import (
    EvidenceDocument,
    ImageBlock,
    ParsedDocument,
    ParsedPage,
    TableBlock,
    TextBlock,
)

OFFICE_EXTENSIONS = frozenset({".docx", ".pptx", ".xlsx", ".xls"})


def load_office_document(path: Path, document: EvidenceDocument) -> ParsedDocument:
    suffix = path.suffix.lower()
    if suffix not in OFFICE_EXTENSIONS:
        raise ValueError(f"unsupported Office format: {suffix}")
    if suffix in {".xlsx", ".xls"}:
        return _load_workbook(path, document)

    markdown = _docling_markdown(path)
    fallback_chain = ["docling"]
    parser = "docling"
    archive_pages = _archive_pages(path, suffix)
    if not markdown:
        markdown = "\n\n".join(archive_pages)
        fallback_chain.append("office_xml")
        parser = "office_xml"
    elif suffix == ".pptx" and archive_pages:
        fallback_chain.append("office_xml_page_map")
    images = _archive_images(path, document)
    page_texts = archive_pages if suffix == ".pptx" and archive_pages else [markdown]
    pages = tuple(ParsedPage(page=index, text=text) for index, text in enumerate(page_texts, start=1))
    blocks = tuple(
        TextBlock(block_id=_id(document, "text", index), page=index, text=text)
        for index, text in enumerate(page_texts, start=1)
        if text
    )
    return ParsedDocument(
        document=document,
        pages=pages,
        text_blocks=blocks,
        images=images,
        parser=parser,
        fallback_chain=tuple(fallback_chain),
    )


def parsed_to_documents(parsed: ParsedDocument) -> list[Document]:
    base = {
        "source": parsed.document.source,
        "filename": parsed.document.filename,
        "document_id": parsed.document.document_id,
        "version": parsed.document.version,
        "tenant_id": parsed.document.tenant_id,
        "owner_user_id": parsed.document.owner_user_id,
        "visibility": parsed.document.visibility,
        "acl_tags": ",".join(parsed.document.acl_tags),
        "parser": parsed.parser,
    }
    documents = [
        Document(
            page_content=block.text,
            metadata={**base, "page": block.page, "sheet": block.sheet or "", "block_id": block.block_id, "modality": "text"},
        )
        for block in parsed.text_blocks
        if block.text.strip()
    ]
    documents.extend(
        Document(
            page_content=table.markdown,
            metadata={**base, "page": table.page, "sheet": table.sheet or "", "table_id": table.table_id, "modality": "table"},
        )
        for table in parsed.tables
        if table.markdown.strip()
    )
    documents.extend(
        Document(
            page_content=(image.ocr_text or image.description or f"Image artifact {image.image_id}"),
            metadata={**base, "page": image.page, "image_id": image.image_id, "modality": "image"},
        )
        for image in parsed.images
    )
    return documents


def _load_workbook(path: Path, document: EvidenceDocument) -> ParsedDocument:
    sheets = _xlsx_rows(path) if path.suffix.lower() == ".xlsx" else _xls_rows(path)
    pages: list[ParsedPage] = []
    tables: list[TableBlock] = []
    for page_number, (sheet, rows) in enumerate(sheets, start=1):
        markdown = _rows_to_markdown(rows)
        pages.append(ParsedPage(page=page_number, sheet=sheet, text=markdown))
        if markdown:
            tables.append(
                TableBlock(
                    table_id=_id(document, "table", page_number),
                    page=page_number,
                    sheet=sheet,
                    markdown=markdown,
                )
            )
    parser = "openpyxl" if path.suffix.lower() == ".xlsx" else "pandas"
    return ParsedDocument(
        document=document,
        pages=tuple(pages),
        tables=tuple(tables),
        parser=parser,
        fallback_chain=(parser,),
    )


def _xlsx_rows(path: Path) -> list[tuple[str, list[list[object]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX ingestion requires the 'office' dependency extra") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [(sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]) for sheet in workbook.worksheets]
    finally:
        workbook.close()


def _xls_rows(path: Path) -> list[tuple[str, list[list[object]]]]:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("XLS ingestion requires the 'office' dependency extra") from exc
    try:
        frames = pd.read_excel(path, sheet_name=None, header=None)
    except ImportError as exc:
        raise RuntimeError("XLS ingestion requires xlrd from the 'office' dependency extra") from exc
    return [(str(name), frame.fillna("").values.tolist()) for name, frame in frames.items()]


def _docling_markdown(path: Path) -> str:
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        return ""
    try:
        return str(DocumentConverter().convert(str(path)).document.export_to_markdown() or "").strip()
    except Exception:
        return ""


def _archive_pages(path: Path, suffix: str) -> list[str]:
    prefix = "word/" if suffix == ".docx" else "ppt/slides/"
    with zipfile.ZipFile(path) as archive:
        names = sorted(name for name in archive.namelist() if name.startswith(prefix) and name.endswith(".xml"))
        texts: list[str] = []
        for name in names:
            root = ElementTree.fromstring(archive.read(name))
            page_text = "\n".join(value.strip() for node in root.iter() if (value := node.text) and value.strip())
            if page_text:
                texts.append(page_text)
        return texts


def _archive_images(path: Path, document: EvidenceDocument) -> tuple[ImageBlock, ...]:
    prefix = "word/media/" if path.suffix.lower() == ".docx" else "ppt/media/"
    with zipfile.ZipFile(path) as archive:
        page_by_name = _ppt_image_pages(archive) if path.suffix.lower() == ".pptx" else {}
        names = sorted(name for name in archive.namelist() if name.startswith(prefix) and not name.endswith("/"))
        return tuple(
            ImageBlock(
                image_id=_id(document, "image", index),
                page=page_by_name.get(Path(name).name, 1),
                filename=Path(name).name,
                media_type=mimetypes.guess_type(name)[0] or "application/octet-stream",
                data=archive.read(name),
            )
            for index, name in enumerate(names, start=1)
        )


def _ppt_image_pages(archive: zipfile.ZipFile) -> dict[str, int]:
    pages: dict[str, int] = {}
    relationship_names = sorted(
        name
        for name in archive.namelist()
        if name.startswith("ppt/slides/_rels/slide") and name.endswith(".xml.rels")
    )
    for relationship_name in relationship_names:
        match = re.search(r"slide(\d+)\.xml\.rels$", relationship_name)
        if not match:
            continue
        page = int(match.group(1))
        root = ElementTree.fromstring(archive.read(relationship_name))
        for node in root.iter():
            target = str(node.attrib.get("Target", "") or "")
            if "/media/" in target or target.startswith("../media/"):
                pages.setdefault(Path(target).name, page)
    return pages


def _rows_to_markdown(rows: list[list[object]]) -> str:
    normalized = [[_cell(value) for value in row] for row in rows]
    normalized = [row for row in normalized if any(cell for cell in row)]
    if not normalized:
        return ""
    width = max(len(row) for row in normalized)
    padded = [row + [""] * (width - len(row)) for row in normalized]
    header = padded[0]
    body = padded[1:]
    return "\n".join(
        ["| " + " | ".join(header) + " |", "| " + " | ".join("---" for _ in header) + " |"]
        + ["| " + " | ".join(row) + " |" for row in body]
    )


def _cell(value: object) -> str:
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip().replace("|", "\\|")


def _id(document: EvidenceDocument, kind: str, index: int) -> str:
    seed = f"{document.document_id}|{document.version}|{kind}|{index}"
    return f"{kind}-{hashlib.sha256(seed.encode('utf-8')).hexdigest()[:20]}"


__all__ = ["OFFICE_EXTENSIONS", "load_office_document", "parsed_to_documents"]
